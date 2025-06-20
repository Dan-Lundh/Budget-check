import fitz  # PyMuPDF
import openpyxl
#from features.gemini.agent import Agent
import pandas as pd
import numpy as np
import numpy as np
from numpy import array
from numpy import argmax
import matplotlib.pyplot as plt


from PySide6.QtGui import QImage, QPixmap, QColor

#agent = Agent()



def pdf_to_text(pdf_file_path):
    """
    Reads an pdf file using PyMuPDF/fitz and converts its content to a text file.

    Args:
        pdf_file_path (str): The path to the pdf file.
        Function return the text.
    """
    doc = fitz.open(pdf_file_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def excel_to_text_openpyxl(excel_file_path, output_text_path=None):
    """
    Reads an Excel file using openpyxl and converts its content to a text file.

    Args:
        excel_file_path (str): The path to the Excel file.
        output_text_path (str, optional): The path to save the output text file.
                                         If None, function return the text.
    """
    try:
        workbook = openpyxl.load_workbook(excel_file_path)

        text_output = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_output += f"--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_values = [str(cell.value) for cell in row]
                text_output += "\t".join(row_values) + "\n"
            text_output += "\n"

        if output_text_path:
            with open(output_text_path, 'w', encoding='utf-8') as text_file:
                text_file.write(text_output)
            print(f"Excel data written to: {output_text_path}")
        else:
            return text_output
        
    except FileNotFoundError:
        print(f"Error: File not found at {excel_file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")



def fact_agent(text_report):
    prompt=f"Present all the numerical values in {text_report} using a csv table"
    respons= agent.talk(prompt=prompt)
    return respons.text

def read_stat_proj(filename):
    df= pd.read_excel(filename,header=0)
    return df

def filter(df,proj_id):
    df2=pd.DataFrame()
    cat_list2=pd.unique(df[df["Project"]==proj_id]["Category"])
    date_list3=pd.unique(df[df["Project"]==proj_id]["Report Date"])
    for category in cat_list2:
        for date in date_list3: 
            df_temp=df[(df["Project"] == proj_id) & 
                     (df["Category"] == category) & 
                     (df["Report Date"] == date)][["Category","Report Date","Man-Hours (Est)","Man-Hours (Actual)"]]
            if not df_temp.empty:
                df2 = pd.concat([df2, df_temp], ignore_index=True)
    return df2, date_list3


def calc_ratio(df):
    df_filtered = df[df["Category"] != "Total Project Est:"].copy()
    df_filtered["Ratio"] = df_filtered["Man-Hours (Actual)"] / df_filtered["Man-Hours (Est)"]
    result_df = df_filtered.pivot(index="Report Date", columns="Category", values="Ratio").reset_index()
    result_df.columns.name = None  # Remove the columns index name
    #result_df = result_df.sort_index(axis=1)
    return result_df


def calc_per_date(df):
    df_filtered = df[df["Category"] != "Total Project Est:"].copy()
    total_per_date = df_filtered.groupby("Report Date")["Man-Hours (Actual)"].transform("sum")
    df_filtered["Actual_vs_DailyTotal"] = df_filtered["Man-Hours (Actual)"] / total_per_date
    pivot_df = df_filtered.pivot(index="Report Date", columns="Category", values="Actual_vs_DailyTotal")
    # Optional: sort the columns alphabetically
    # pivot_df = pivot_df.sort_index(axis=1)
    return pivot_df


def norm_image(pivot_df):
    pivot_df_filled = pivot_df.fillna(0)
    if 'Report Date' in pivot_df_filled.columns:
        pivot_df_filled = pivot_df_filled.drop(columns='Report Date')
    
    image_vectors = pivot_df_filled.to_numpy()
    return image_vectors

def join_data(df1, df2):
    array1=np.array(df1)
    array2=np.array(df2)
    combined = np.hstack((array1, array2))
    return combined 

def arrayimage(img):
    gray=np.array(img)
    rgb = np.stack([gray]*3, axis=-1)
    return np.array(rgb/255)

def toOneHot(data):
    code=[]
    for x in data:
        if x=='normal':
            code.append(0)
        else:
            code.append(1)
    return np.array(code)

def numpy_to_qimage(np_array):
    """
    Converts a NumPy array (representing an image) to a QImage.
    Supports grayscale (2D) and RGB/RGBA (3D) uint8 arrays.
    """
    height, width = np_array.shape[0], np_array.shape[1]

    if np_array.ndim == 2:
        # Grayscale image (H, W)
        if np_array.dtype != np.uint8:
            # Scale to 0-255 if not already uint8
            np_array = (np_array / np_array.max() * 255).astype(np.uint8)

        # QImage.Format_Grayscale8 requires a 1-byte per pixel image
        # Ensure it's C-contiguous for QImage.data to work directly
        img_data = np.ascontiguousarray(np_array)
        q_image = QImage(
            img_data.data,
            width,
            height,
            width,  # bytesPerLine for grayscale is just width
            QImage.Format_Grayscale8
        )
        # For grayscale, you often need to set a color table to map pixel values to grayscale colors
        # (0-255 map to black-white)
        q_image.setColorTable([QColor(i, i, i).rgb() for i in range(256)])
        return q_image

    elif np_array.ndim == 3:
        # Color image (H, W, C)
        if np_array.dtype != np.uint8:
            #np_array = (np_array / np_array.max() * 255).astype(np.uint8)
            np_array = (np_array * 255).astype(np.uint8)

        channels = np_array.shape[2]
        bytes_per_line = channels * width

        if channels == 3:
            # RGB image (H, W, 3) -> QImage.Format_RGB888
            # Ensure it's C-contiguous and in RGB order
            img_data = np.ascontiguousarray(np_array)
            q_image = QImage(
                img_data.data,
                width,
                height,
                bytes_per_line,
                QImage.Format_RGB888
            )
            return q_image
        elif channels == 4:
            # RGBA image (H, W, 4) -> QImage.Format_RGBA8888 or Format_ARGB32
            # QImage.Format_RGBA8888 expects data as R-G-B-A
            # Ensure it's C-contiguous
            img_data = np.ascontiguousarray(np_array)
            q_image = QImage(
                img_data.data,
                width,
                height,
                bytes_per_line,
                QImage.Format_RGBA8888 # Or Format_ARGB32 for (A,R,G,B)
            )
            return q_image
        else:
            raise ValueError(f"Unsupported number of channels: {channels}. Expected 1 (grayscale), 3 (RGB), or 4 (RGBA).")
    else:
        raise ValueError(f"Unsupported array dimensions: {np_array.ndim}. Expected 2D or 3D.")


def rgb_to_grayscale_average(rgb_image):
    """Converts an RGB image (NumPy array) to grayscale using the average method."""
    if rgb_image.ndim != 3 or rgb_image.shape[2] != 3:
        raise ValueError("Input image must be a 3D RGB array (H, W, 3).")
    grayscale_image = np.mean(rgb_image, axis=2, dtype=np.uint8)
    return grayscale_image

def rgb_to_grayscale_weighted(rgb_image):
    """Converts an RGB image to grayscale using a weighted sum (more perceptually accurate)."""
    if rgb_image.ndim != 3 or rgb_image.shape[2] != 3:
        raise ValueError("Input image must be a 3D RGB array (H, W, 3).")
    # Ensure input is float for calculation to avoid overflow if intermediate sum > 255
    # Then round and cast to uint8
    grayscale_image = np.round(
        0.2126 * rgb_image[:, :, 0].astype(np.float32) +
        0.7152 * rgb_image[:, :, 1].astype(np.float32) +
        0.0722 * rgb_image[:, :, 2].astype(np.float32)
    ).astype(np.uint8)
    return grayscale_image